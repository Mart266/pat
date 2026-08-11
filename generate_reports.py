#!/usr/bin/env python3
"""
Test Tag Melbourne — test and tag report generator.

Takes the ZIP (or bare CSV) exported by the PAT Logger and produces:
  1. Certificate of Conformance      (PDF, 1 page)
  2. Concise Test Report             (PDF, grouped by area)
  3. Detailed Test Report            (PDF, per-item test detail)
  4. Fail Report                     (PDF, with photos, only if failures exist)
  5. Detailed Test Register          (XLSX)

Usage:
    python generate_reports.py <export.zip | export.csv> [output_dir]
"""

import csv
import io
import os
import sys
import zipfile
from collections import OrderedDict, defaultdict
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (BaseDocTemplate, Frame, Image, KeepTogether,
                                PageBreak, PageTemplate, Paragraph, Spacer,
                                Table, TableStyle)

# ---------------------------------------------------------------- configuration

# ---------------------------------------------------------------- configuration
#
# Business details are read from report_config.json sitting beside this script.
# That file is deliberately not committed, so nothing identifying lives in the
# repository. Copy report_config.example.json and fill it in.

CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "report_config.json")

DEFAULTS = {
    "company": {"name": "", "abn": "", "phone": "", "web": "", "address": ""},
    "instrument": {"description": "", "calibration": ""},
}


def load_config():
    cfg = {k: dict(v) for k, v in DEFAULTS.items()}
    if os.path.exists(CONFIG_PATH):
        import json as _json
        with open(CONFIG_PATH, encoding="utf-8") as f:
            data = _json.load(f)
        for section in cfg:
            cfg[section].update(data.get(section, {}))
    else:
        print("  note: no report_config.json found; the letterhead will be blank")
    return cfg


_CFG = load_config()
COMPANY = _CFG["company"]
INSTRUMENT = _CFG["instrument"]

STANDARD = "AS/NZS 3760:2022"
STANDARD_LONG = "AS/NZS 3760:2022 In-service safety inspection and testing of electrical equipment"

INSPECTION_ITEMS = [
    "Cover / Guards", "Flexible cord", "Plugs", "Controls", "Safety Devices",
    "Accessories", "Max Load Label", "Ventilation", "Outlet Sockets", "Connectors",
]

# ---------------------------------------------------------------------- palette

INK = colors.HexColor("#18181b")
MUTED = colors.HexColor("#71717a")
LINE = colors.HexColor("#d4d4d8")
FAINT = colors.HexColor("#f4f4f5")
BRAND_RED = colors.HexColor("#D2030F")
BAND = colors.HexColor("#0a0a0a")
PASS_C = colors.HexColor("#15803d")
FAIL_C = colors.HexColor("#b91c1c")

MARGIN = 12 * mm
HEADER_H = 34 * mm
FOOTER_H = 14 * mm


# ------------------------------------------------------------------- data input

def load_export(path):
    """Return (rows, photos) where photos maps filename -> bytes."""
    photos = {}
    if path.lower().endswith(".zip"):
        with zipfile.ZipFile(path) as z:
            csv_names = [n for n in z.namelist() if n.lower().endswith(".csv")]
            if not csv_names:
                raise SystemExit("No CSV found inside the zip.")
            text = z.read(csv_names[0]).decode("utf-8-sig")
            for n in z.namelist():
                if n.lower().endswith((".jpg", ".jpeg", ".png")):
                    photos[os.path.basename(n)] = z.read(n)
    else:
        with open(path, encoding="utf-8-sig") as f:
            text = f.read()

    rows = list(csv.DictReader(io.StringIO(text)))
    if not rows:
        raise SystemExit("The export contains no test records.")
    return rows, photos


def parse_iso(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def fmt_date(dt):
    if not dt:
        return ""
    day = dt.day
    if 11 <= day <= 13:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return "%d%s %s %d" % (day, suf, dt.strftime("%b"), dt.year)


def job_context(rows):
    """Pull the job-level facts, flagging inconsistency rather than hiding it."""
    def distinct(key):
        return sorted({(r.get(key) or "").strip() for r in rows} - {""})

    clients, sites, refs, techs = (distinct(k) for k in
                                  ("Client", "Site", "Job Reference", "Technician"))
    dates = sorted(d for d in (parse_iso(r.get("Test Date ISO", "")) for r in rows) if d)

    warnings = []
    for label, values in (("client", clients), ("site", sites), ("job reference", refs)):
        if len(values) > 1:
            warnings.append("More than one %s in this export: %s" % (label, ", ".join(values)))

    return {
        "client": clients[0] if clients else "",
        "site": sites[0] if sites else "",
        "ref": refs[0] if refs else "",
        "technicians": techs,
        "start": dates[0] if dates else None,
        "end": dates[-1] if dates else None,
        "warnings": warnings,
    }


def is_fail(row):
    return (row.get("Result") or "").strip().upper() == "FAIL"


def by_area(rows):
    groups = OrderedDict()
    for r in rows:
        area = (r.get("Location") or "Unspecified").strip() or "Unspecified"
        groups.setdefault(area, []).append(r)
    for area in groups:
        groups[area].sort(key=lambda r: r.get("Asset ID", ""))
    return groups


# ------------------------------------------------------------------- pdf chrome

class ReportDoc(BaseDocTemplate):
    """A4 with the company header on every page and a three-part footer."""

    def __init__(self, path, report_name, ctx, **kw):
        self.report_name = report_name
        self.ctx = ctx
        BaseDocTemplate.__init__(self, path, pagesize=A4,
                                 leftMargin=MARGIN, rightMargin=MARGIN,
                                 topMargin=MARGIN + HEADER_H,
                                 bottomMargin=MARGIN + FOOTER_H,
                                 title="%s - %s" % (COMPANY["name"] or "Test report", report_name),
                                 author=COMPANY["name"] or "", **kw)
        frame = Frame(self.leftMargin, self.bottomMargin,
                      self.width, self.height, id="body",
                      leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
        self.addPageTemplates([PageTemplate(id="main", frames=[frame],
                                            onPage=self._decorate)])

    def _decorate(self, canv, doc):
        w, h = A4
        canv.saveState()

        # --- letterhead band, so the logo sits on the dark ground it needs
        band_h = 26 * mm
        band_top = h - MARGIN
        canv.setFillColor(BAND)
        canv.rect(MARGIN, band_top - band_h, w - 2 * MARGIN, band_h, stroke=0, fill=1)
        canv.setFillColor(BRAND_RED)
        canv.rect(MARGIN, band_top - band_h - 1.4 * mm, w - 2 * MARGIN, 1.4 * mm,
                  stroke=0, fill=1)

        logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png")
        if os.path.exists(logo):
            img = ImageReader(logo)
            iw, ih = img.getSize()
            # fit inside the band with breathing room on both axes
            max_w, max_h = 44 * mm, band_h - 9 * mm
            scale = min(max_w / float(iw), max_h / float(ih))
            lw, lh = iw * scale, ih * scale
            canv.drawImage(logo, MARGIN + 6 * mm,
                           band_top - band_h / 2.0 - lh / 2.0, width=lw, height=lh,
                           mask="auto")
        else:
            canv.setFillColor(colors.white)
            canv.setFont("Helvetica-Bold", 15)
            canv.drawString(MARGIN + 6 * mm, band_top - 12 * mm,
                            (COMPANY["name"] or "").upper())

        right = w - MARGIN - 6 * mm
        lines = []
        if COMPANY.get("phone"):
            lines.append(("Helvetica-Bold", 7.4, colors.white, COMPANY["phone"]))
        for key, prefix in (("web", ""), ("address", ""), ("abn", "ABN ")):
            if COMPANY.get(key):
                lines.append(("Helvetica", 7.2, colors.HexColor("#c8c8ce"),
                              prefix + COMPANY[key]))
        y_line = band_top - 7 * mm
        for font, size, colour, text in lines:
            canv.setFont(font, size)
            canv.setFillColor(colour)
            canv.drawRightString(right, y_line, text)
            y_line -= 3.6 * mm

        # --- footer
        y = MARGIN + 5 * mm
        canv.setStrokeColor(LINE)
        canv.setLineWidth(0.5)
        canv.line(MARGIN, y + 5 * mm, w - MARGIN, y + 5 * mm)
        canv.setFont("Helvetica", 7)
        canv.setFillColor(MUTED)
        canv.drawString(MARGIN, y, self.report_name)
        canv.drawCentredString(w / 2.0, y, "Page %d" % doc.page)
        tail = " | ".join(x for x in (self.ctx.get("client"), self.ctx.get("site"),
                                     self.ctx.get("ref")) if x)
        canv.drawRightString(w - MARGIN, y, tail)
        canv.restoreState()


def styles():
    s = getSampleStyleSheet()
    return {
        "h1": ParagraphStyle("h1", parent=s["Normal"], fontName="Helvetica-Bold",
                             fontSize=15, leading=19, textColor=INK, spaceAfter=2 * mm),
        "h2": ParagraphStyle("h2", parent=s["Normal"], fontName="Helvetica-Bold",
                             fontSize=9, leading=12, textColor=INK,
                             spaceBefore=4 * mm, spaceAfter=1.5 * mm),
        "body": ParagraphStyle("body", parent=s["Normal"], fontName="Helvetica",
                               fontSize=8.5, leading=12, textColor=INK),
        "small": ParagraphStyle("small", parent=s["Normal"], fontName="Helvetica",
                                fontSize=7.5, leading=10.5, textColor=MUTED),
        "cell": ParagraphStyle("cell", parent=s["Normal"], fontName="Helvetica",
                               fontSize=7.5, leading=9.5, textColor=INK),
        "centre": ParagraphStyle("centre", parent=s["Normal"], fontName="Helvetica",
                                 fontSize=8.5, leading=12, alignment=TA_CENTER,
                                 textColor=INK),
    }


def meta_block(ctx, rows, st, width):
    """Client / job / summary strip used at the top of each report."""
    total = len(rows)
    fails = sum(1 for r in rows if is_fail(r))
    period = fmt_date(ctx["start"])
    if ctx["end"] and ctx["end"] != ctx["start"]:
        period += " to " + fmt_date(ctx["end"])

    def block(label, lines):
        out = ['<font size="6.5" color="#71717a">%s</font>' % label.upper()]
        out += ['<font size="8.5">%s</font>' % ln for ln in lines if ln]
        return Paragraph("<br/>".join(out), st["body"])

    data = [[
        block("Client", [ctx["client"] or "\u2014", ctx["site"]]),
        block("Job", [ctx["ref"] or "\u2014", period]),
        block("Summary", ["%d tests" % total,
                          "%d %s" % (fails, "failure" if fails == 1 else "failures")]),
    ]]
    t = Table(data, colWidths=[width * 0.42, width * 0.34, width * 0.24])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
        ("RIGHTPADDING", (-1, 0), (-1, 0), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4 * mm),
    ]))
    return t


def instrument_note(st):
    bits = []
    if INSTRUMENT.get("description"):
        bits.append("Testing carried out with %s." % INSTRUMENT["description"])
    if INSTRUMENT.get("calibration"):
        bits.append("Instrument calibration: %s." % INSTRUMENT["calibration"])
    bits.append("Tests performed to the requirements of %s." % STANDARD_LONG)
    return Paragraph(" ".join(bits), st["small"])


# --------------------------------------------------------- 1. certificate

def certificate(path, rows, ctx):
    st = styles()
    doc = ReportDoc(path, "Certificate of Conformance", ctx)
    W = doc.width
    story = []

    total = len(rows)
    fails = sum(1 for r in rows if is_fail(r))
    passes = total - fails

    story.append(Paragraph("Certificate of Conformance", st["h1"]))
    story.append(Spacer(1, 2 * mm))

    where = ", ".join(x for x in (ctx["client"], ctx["site"]) if x) or "The client"
    period = fmt_date(ctx["start"])
    if ctx["end"] and ctx["end"] != ctx["start"]:
        period += " to " + fmt_date(ctx["end"])

    story.append(Paragraph(
        "%s presented the following equipment for in-service electrical safety "
        "inspection and testing during the period <b>%s</b>." % (where, period),
        st["body"]))
    story.append(Spacer(1, 6 * mm))

    def section(title, lines):
        story.append(Paragraph(title, st["h2"]))
        data = [[Paragraph(text, st["body"]),
                 Paragraph('<para align="right">%s</para>' % value, st["body"])]
                for text, value in lines]
        t = Table(data, colWidths=[W - 22 * mm, 22 * mm])
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.4, LINE),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5 * mm),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5 * mm),
            ("LEFTPADDING", (0, 0), (0, -1), 0),
            ("RIGHTPADDING", (-1, 0), (-1, -1), 0),
        ]))
        story.append(t)

    section("Electrical appliances and leads", [
        ("Items that satisfied the thresholds detailed in %s" % STANDARD, str(passes)),
        ("Items that did not satisfy the thresholds detailed in %s" % STANDARD, str(fails)),
    ])

    section("Not within the scope of this inspection", [
        ("RCDs \u2014 trip-time testing not performed", "N/A"),
        ("Microwaves \u2014 leakage testing to AS/NZS 60335.2.25 not performed", "N/A"),
        ("Emergency and exit lighting \u2014 AS/NZS 2293.2", "N/A"),
    ])

    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        "%s has inspected and tested the equipment found at the above site and recorded "
        "the results in the accompanying reports. The information above is accurate at the "
        "time of issue. Please advise %s immediately of any error or omission."
        % (COMPANY["name"] or "This business",
           COMPANY["name"] or "this business"), st["body"]))

    if fails:
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(
            "<b>%d item%s failed and %s been removed from service.</b> Refer to the Fail "
            "Report for the defects recorded and the action taken."
            % (fails, "" if fails == 1 else "s", "has" if fails == 1 else "have"),
            st["body"]))

    story.append(Spacer(1, 12 * mm))
    tech = ", ".join(ctx["technicians"]) or "\u2014"
    sig = Table([
        [Paragraph('<font size="6.5" color="#71717a">TESTED BY</font><br/>'
                   '<font size="9">%s</font>' % tech, st["body"]),
         Paragraph('<font size="6.5" color="#71717a">DATE OF ISSUE</font><br/>'
                   '<font size="9">%s</font>' % fmt_date(datetime.now()), st["body"]),
         Paragraph('<font size="6.5" color="#71717a">SIGNED</font>', st["body"])],
    ], colWidths=[W * 0.36, W * 0.28, W * 0.36], rowHeights=[16 * mm])
    sig.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, INK),
        ("LEFTPADDING", (0, 0), (0, -1), 0),
    ]))
    story.append(sig)

    story.append(Spacer(1, 6 * mm))
    story.append(instrument_note(st))

    doc.build(story)


# --------------------------------------------------------- 2. concise report

def concise(path, rows, ctx):
    st = styles()
    doc = ReportDoc(path, "Test and Tag \u2014 Concise Test Report", ctx)
    W = doc.width
    story = [Paragraph("Concise Test Report", st["h1"]),
             meta_block(ctx, rows, st, W)]

    cols = [W * 0.15, W * 0.29, W * 0.16, W * 0.12, W * 0.16, W * 0.12]
    head = ["ASSET ID", "ASSET", "TECHNICIAN", "TEST DATE", "NEXT TEST", "RESULT"]

    for area, items in by_area(rows).items():
        a_fail = sum(1 for r in items if is_fail(r))
        story.append(Paragraph(
            '%s<font color="#71717a" size="7.5">&nbsp;&nbsp;%d of %d passed</font>'
            % (area.upper(), len(items) - a_fail, len(items)), st["h2"]))

        data = [[Paragraph('<font size="6.5" color="#71717a">%s</font>' % h, st["cell"])
                 for h in head]]
        styling = [
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, INK),
            ("LINEBELOW", (0, 1), (-1, -2), 0.3, FAINT),
            ("TOPPADDING", (0, 0), (-1, -1), 1.6 * mm),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.6 * mm),
            ("LEFTPADDING", (0, 0), (0, -1), 0),
            ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
            ("RIGHTPADDING", (-1, 0), (-1, -1), 0),
        ]

        for i, r in enumerate(items, start=1):
            failed = is_fail(r)
            result = "Fail" if failed else "Pass"
            data.append([
                Paragraph('<font face="Courier">%s</font>' % r.get("Asset ID", ""), st["cell"]),
                Paragraph(r.get("Description", ""), st["cell"]),
                Paragraph(r.get("Technician", ""), st["cell"]),
                Paragraph(r.get("Test Date", ""), st["cell"]),
                Paragraph(r.get("Next Test Date", "") or "\u2014", st["cell"]),
                Paragraph('<b>%s</b>' % result, st["cell"]),
            ])
            styling.append(("TEXTCOLOR", (-1, i), (-1, i), FAIL_C if failed else PASS_C))

        t = Table(data, colWidths=cols, repeatRows=1)
        t.setStyle(TableStyle(styling))
        story.append(t)

    story.append(Spacer(1, 6 * mm))
    story.append(instrument_note(st))
    doc.build(story)


# --------------------------------------------------------- 3. detailed report

def detailed_report(path, rows, ctx):
    st = styles()
    doc = ReportDoc(path, "Test and Tag — Detailed Test Report", ctx)
    W = doc.width
    story = [Paragraph("Detailed Test Report", st["h1"]),
             meta_block(ctx, rows, st, W)]

    for area, items in by_area(rows).items():
        story.append(Paragraph(area.upper(), st["h2"]))
        for r in items:
            failed = is_fail(r)
            asset = r.get("Description", "")
            tag = r.get("Asset ID", "")

            hdr = Table([[
                Paragraph('<b>%s</b>&nbsp;&nbsp;<font size="7.5" color="#71717a" face="Courier">%s</font>'
                          % (asset, tag), st["body"]),
                Paragraph('<para align="right"><b><font color="%s">%s</font></b></para>'
                          % ("#b91c1c" if failed else "#15803d", "FAIL" if failed else "PASS"), st["body"]),
            ]], colWidths=[W * 0.78, W * 0.22])
            hdr.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5 * mm),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, INK),
            ]))

            details = [
                ("Test type", r.get("Test Type", "")),
                ("Location", r.get("Location", "")),
                ("Test date", r.get("Test Date", "")),
                ("Next test", r.get("Next Test Date", "")),
                ("Polarity", r.get("Polarity", "")),
            ]
            if (r.get("Outlets") or "").strip():
                details.append(("Outlets tested", r.get("Outlets", "")))
            if (r.get("Leakage mW/cm2") or "").strip():
                details.append(("Microwave leakage", "%s mW/cm²" % r.get("Leakage mW/cm2", "")))

            sample = (r.get("Sample Value") or r.get("Earth Continuity Value (Ω)") or "").strip()
            if sample:
                vals = [v.strip() for v in sample.split("|") if v.strip()]
                if len(vals) > 1:
                    for n, v in enumerate(vals, start=1):
                        details.append(("Outlet %d sample value" % n, v))
                else:
                    details.append(("Sample value", sample))

            ddata = [[Paragraph('<font size="6.5" color="#71717a">TEST DETAIL</font>', st["cell"]),
                      Paragraph('<para align="right"><font size="6.5" color="#71717a">VALUE</font></para>', st["cell"])]]
            for label, value in details:
                ddata.append([Paragraph(label, st["cell"]),
                              Paragraph('<para align="right">%s</para>' % (value or "—"), st["cell"])])
            dt = Table(ddata, colWidths=[W * 0.30, W * 0.20], hAlign="LEFT")
            dt.setStyle(TableStyle([
                ("LINEBELOW", (0, 0), (-1, 0), 0.4, LINE),
                ("LINEBELOW", (0, 1), (-1, -2), 0.3, FAINT),
                ("TOPPADDING", (0, 0), (-1, -1), 1.1 * mm),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.1 * mm),
                ("LEFTPADDING", (0, 0), (0, -1), 0),
                ("RIGHTPADDING", (-1, 0), (-1, -1), 0),
            ]))

            failed_items = [i.strip() for i in (r.get("Failed Inspection Items") or "").split(";") if i.strip()]
            idata = [[Paragraph('<font size="6.5" color="#71717a">INSPECTION ITEM</font>', st["cell"]),
                      Paragraph('<para align="right"><font size="6.5" color="#71717a">RESULT</font></para>', st["cell"])]]
            istyle = [
                ("LINEBELOW", (0, 0), (-1, 0), 0.4, LINE),
                ("LINEBELOW", (0, 1), (-1, -2), 0.3, FAINT),
                ("TOPPADDING", (0, 0), (-1, -1), 1.1 * mm),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.1 * mm),
                ("LEFTPADDING", (0, 0), (0, -1), 0),
                ("RIGHTPADDING", (-1, 0), (-1, -1), 0),
            ]
            for i, item in enumerate(INSPECTION_ITEMS, start=1):
                value = (r.get(item) or "").strip() or ("Fail" if item in failed_items else "Pass")
                bad = value.lower() == "fail" or item in failed_items
                idata.append([Paragraph(item, st["cell"]),
                              Paragraph('<para align="right"><b>%s</b></para>' % value, st["cell"])])
                istyle.append(("TEXTCOLOR", (-1, i), (-1, i), FAIL_C if bad else PASS_C))
            it = Table(idata, colWidths=[W * 0.36, W * 0.14], hAlign="LEFT")
            it.setStyle(TableStyle(istyle))

            body = Table([[dt, it]], colWidths=[W * 0.5, W * 0.5], hAlign="LEFT")
            body.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (-1, 0), (-1, 0), 0),
                ("LEFTPADDING", (-1, 0), (-1, 0), 5 * mm),
                ("TOPPADDING", (0, 0), (-1, -1), 2 * mm),
            ]))

            story.append(KeepTogether([hdr, body, Spacer(1, 5 * mm)]))

    story.append(instrument_note(st))
    doc.build(story)


# --------------------------------------------------------- 4. fail report

def fail_report(path, rows, ctx, photos):
    failures = [r for r in rows if is_fail(r)]
    if not failures:
        return False

    st = styles()
    doc = ReportDoc(path, "Test and Tag \u2014 Fail Report", ctx)
    W = doc.width
    story = [Paragraph("Fail Report", st["h1"]),
             meta_block(ctx, rows, st, W)]

    story.append(Paragraph(
        "%d item%s failed inspection or testing and %s been removed from service."
        % (len(failures), "" if len(failures) == 1 else "s",
           "has" if len(failures) == 1 else "have"), st["body"]))

    for area, items in by_area(failures).items():
        pending_heading = Paragraph(area.upper(), st["h2"])

        for r in items:
            block = []
            if pending_heading is not None:
                block.append(pending_heading)
                pending_heading = None
            asset = r.get("Description", "")
            tag = r.get("Asset ID", "")
            hdr = Table([[
                Paragraph('<b>%s</b>&nbsp;&nbsp;<font size="7.5" color="#71717a" '
                          'face="Courier">%s</font>' % (asset, tag), st["body"]),
                Paragraph('<para align="right"><b><font color="#b91c1c">FAIL</font></b>'
                          '</para>', st["body"]),
            ]], colWidths=[W * 0.78, W * 0.22])
            hdr.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5 * mm),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, INK),
            ]))
            block.append(hdr)

            # visual inspection results, item by item
            failed_items = [i.strip() for i in
                            (r.get("Failed Inspection Items") or "").split(";") if i.strip()]
            insp = [[Paragraph('<font size="6.5" color="#71717a">INSPECTION ITEM</font>',
                               st["cell"]),
                     Paragraph('<font size="6.5" color="#71717a">RESULT</font>', st["cell"])]]
            insp_style = [
                ("LINEBELOW", (0, 0), (-1, 0), 0.4, LINE),
                ("LINEBELOW", (0, 1), (-1, -2), 0.3, FAINT),
                ("TOPPADDING", (0, 0), (-1, -1), 1.2 * mm),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.2 * mm),
                ("LEFTPADDING", (0, 0), (0, -1), 0),
                ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
                ("RIGHTPADDING", (-1, 0), (-1, -1), 0),
            ]
            for i, item in enumerate(INSPECTION_ITEMS, start=1):
                value = (r.get(item) or "").strip() or "\u2014"
                bad = value.lower() == "fail" or item in failed_items
                insp.append([Paragraph(item, st["cell"]),
                             Paragraph("<b>%s</b>" % value, st["cell"])])
                insp_style.append(("TEXTCOLOR", (-1, i), (-1, i),
                                   FAIL_C if bad else (PASS_C if value.lower() == "pass" else MUTED)))

            half = W * 0.5
            it = Table(insp, colWidths=[half * 0.72, half * 0.22], hAlign="LEFT")
            it.setStyle(TableStyle(insp_style))

            # follow-up detail
            detail = []
            for label, key in (("Reason for failure", "Fail Reason"),
                               ("Action taken", "Action Taken"),
                               ("Reported to", "Reported To"),
                               ("Notes", "Notes")):
                value = (r.get(key) or "").strip()
                if value:
                    detail.append('<font size="6.5" color="#71717a">%s</font><br/>%s'
                                  % (label.upper(), value))
            if not detail:
                detail.append('<font size="6.5" color="#71717a">FOLLOW-UP</font><br/>'
                              '<font color="#b91c1c">Not recorded</font>')

            side = Paragraph("<br/><br/>".join(detail), st["cell"])

            body = Table([[it, side]], colWidths=[W * 0.5, W * 0.5], hAlign="LEFT")
            body.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (-1, 0), (-1, 0), 0),
                ("LEFTPADDING", (-1, 0), (-1, 0), 6 * mm),
                ("TOPPADDING", (0, 0), (-1, -1), 2 * mm),
            ]))
            block.append(body)

            # photographs
            shots = sorted(n for n in photos
                           if tag and n.startswith(tag.replace("/", "_") + "-"))
            if shots:
                thumbs = []
                for name in shots[:4]:
                    try:
                        img = ImageReader(io.BytesIO(photos[name]))
                        iw, ih = img.getSize()
                        w = 38 * mm
                        thumbs.append(Image(io.BytesIO(photos[name]),
                                            width=w, height=w * ih / float(iw)))
                    except Exception:
                        continue
                if thumbs:
                    block.append(Spacer(1, 2 * mm))
                    block.append(Paragraph(
                        '<font size="6.5" color="#71717a">PHOTOGRAPHS</font>', st["cell"]))
                    block.append(Spacer(1, 1.5 * mm))
                    grid = Table([thumbs], colWidths=[42 * mm] * len(thumbs),
                                 hAlign="LEFT")
                    grid.setStyle(TableStyle([
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (0, 0), 0),
                    ]))
                    block.append(grid)

            block.append(Spacer(1, 6 * mm))
            story.append(KeepTogether(block))

    story.append(instrument_note(st))
    doc.build(story)
    return True


# --------------------------------------------------------- 5. xlsx register

def register(path, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Register"

    headers = list(rows[0].keys())
    ws.append(headers)

    head_font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="18181B")
    thin = Side(style="thin", color="D4D4D8")

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    body_font = Font(name="Arial", size=9)
    fail_font = Font(name="Arial", size=9, bold=True, color="B91C1C")
    pass_font = Font(name="Arial", size=9, color="15803D")
    result_col = headers.index("Result") + 1 if "Result" in headers else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = body_font
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
        if result_col:
            rc = row[result_col - 1]
            value = str(rc.value or "").upper()
            if value == "FAIL":
                rc.font = fail_font
            elif value == "PASS":
                rc.font = pass_font

    for i, h in enumerate(headers, start=1):
        longest = max([len(h)] + [len(str(r.get(h, "") or "")) for r in rows])
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 9), 42)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), ws.max_row)
    wb.save(path)


# ------------------------------------------------------------------------- main

def main():
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)

    src = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else "reports"
    os.makedirs(out_dir, exist_ok=True)

    rows, photos = load_export(src)
    ctx = job_context(rows)

    for w in ctx["warnings"]:
        print("  warning: %s" % w)

    base = ctx["ref"] or ("%s-%s" % (ctx["client"] or "Job", ctx["site"] or "Site"))
    base = "".join(ch if (ch.isalnum() or ch in "-_") else "-" for ch in base).strip("-")

    produced = []

    p = os.path.join(out_dir, "%s - Certificate of Conformance.pdf" % base)
    certificate(p, rows, ctx)
    produced.append(p)

    p = os.path.join(out_dir, "%s - Concise Test Report.pdf" % base)
    concise(p, rows, ctx)
    produced.append(p)

    p = os.path.join(out_dir, "%s - Detailed Test Report.pdf" % base)
    detailed_report(p, rows, ctx)
    produced.append(p)

    p = os.path.join(out_dir, "%s - Fail Report.pdf" % base)
    if fail_report(p, rows, ctx, photos):
        produced.append(p)
    else:
        print("  no failures recorded, so no Fail Report was produced")

    p = os.path.join(out_dir, "%s - Test Register.xlsx" % base)
    register(p, rows)
    produced.append(p)

    print("\n%d records | %d failures | %d photos"
          % (len(rows), sum(1 for r in rows if is_fail(r)), len(photos)))
    for f in produced:
        print("  " + f)


if __name__ == "__main__":
    main()
